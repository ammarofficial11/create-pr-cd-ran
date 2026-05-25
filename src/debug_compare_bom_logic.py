import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from pathlib import Path


CALCULATED_FILE = Path("output/debug_calculated.xlsx")
BOM_CHECK_FILE = Path("BOMchecklogic.xlsx")
OUTPUT_FILE = Path("output/BOMchecklogic_Compared.xlsx")

CALC_FEEDER_COLUMN = "Feeder 1/2"
BOM_FEEDER_COLUMN = "Loose Jumper-1/2"
CALC_SITE_CODE_COLUMN = "Site Code"
BOM_SITE_CODE_COLUMN = "Site Code"


def load_files() -> tuple[pd.DataFrame, pd.DataFrame, str]:
    global BOM_SITE_CODE_COLUMN

    if not CALCULATED_FILE.exists():
        raise FileNotFoundError(f"Calculated file not found: {CALCULATED_FILE}")
    if not BOM_CHECK_FILE.exists():
        raise FileNotFoundError(f"BOM check file not found: {BOM_CHECK_FILE}")

    calc_df = pd.read_excel(CALCULATED_FILE, sheet_name=0)
    bom_df = pd.read_excel(BOM_CHECK_FILE, sheet_name=0)

    # normalize all column names: strip whitespace and newlines
    calc_df.columns = [
        str(c).strip().replace("\n", " ")
        for c in calc_df.columns
    ]
    bom_df.columns = [
        str(c).strip().replace("\n", " ")
        for c in bom_df.columns
    ]

    print("Calculated columns:")
    print(calc_df.columns.tolist())
    print()
    print("BOM columns:")
    print(bom_df.columns.tolist())
    print()

    # detect correct site code column for BOM
    possible_site_cols = [
        "Site Code",
        "*Customer Site Code",
    ]

    for col in possible_site_cols:
        if col in bom_df.columns:
            BOM_SITE_CODE_COLUMN = col
            break

    print(f"Using BOM Site Code column: {BOM_SITE_CODE_COLUMN}")
    print()

    calc_df[CALC_SITE_CODE_COLUMN] = calc_df[CALC_SITE_CODE_COLUMN].astype(str).str.strip()
    bom_df[BOM_SITE_CODE_COLUMN] = bom_df[BOM_SITE_CODE_COLUMN].astype(str).str.strip()

    return calc_df, bom_df, BOM_SITE_CODE_COLUMN


def compare_values(calc_value, manual_value):
    calc_value = float(calc_value) if calc_value not in [None, "", 0] else 0
    manual_value = float(manual_value) if manual_value not in [None, "", 0] else 0

    if calc_value == manual_value:
        return "MATCH", "✅ MATCH"

    if manual_value > 0 and calc_value == 0:
        return "CALC_ZERO", (
            "🟥 CALC_ZERO - "
            "Calculated returned 0 while manual formula returned value. "
            "Check threshold logic, connector qty, or ECC mapping coverage."
        )

    if calc_value > 0 and manual_value == 0:
        return "MANUAL_ZERO", (
            "🟧 MANUAL_ZERO - "
            "Calculated value exists but manual formula returned 0. "
            "Check manual Excel formula or source columns."
        )

    if abs(calc_value - manual_value) == 1:
        return "ROUNDING_DIFF", (
            "⚠️ ROUNDING_DIFF - "
            "Possible rounding or connector-pair division mismatch."
        )

    return "MISMATCH", (
        "❌ MISMATCH - "
        "Mismatch detected. Check feeder length, connector sum, threshold logic, "
        "helper ItemKey calculation, or ECC mapping coverage."
    )


def build_comparison_df(calc_df: pd.DataFrame, bom_df: pd.DataFrame, bom_site_col: str) -> pd.DataFrame:
    METADATA_COLS = {
        "Site Code",
        "Site Name",
        "Region",
        "DU Code",
        "Subcon TI",
        "Contract Number",
        "Purchasing Area",
        CALC_SITE_CODE_COLUMN,
        bom_site_col,
    }

    # Find common ItemKey columns (exclude metadata)
    calc_item_cols = set(calc_df.columns) - METADATA_COLS
    bom_item_cols = set(bom_df.columns) - METADATA_COLS
    common_cols = sorted(calc_item_cols & bom_item_cols)

    print(f"Common ItemKey columns to compare: {common_cols}")
    print()

    bom_df = bom_df.copy()

    # Add mismatch tracking columns
    bom_df["Mismatch_Columns"] = ""
    bom_df["Mismatch_Summary"] = ""
    bom_df["Total_Mismatch_Count"] = 0
    bom_df["Compare_Status"] = ""
    bom_df["Compare_Remark"] = ""

    for idx, row in bom_df.iterrows():
        site_code = row[bom_site_col]

        calc_row = calc_df[calc_df[CALC_SITE_CODE_COLUMN] == site_code]

        mismatches = []
        mismatch_details = []

        if len(calc_row) > 0:
            calc_row_data = calc_row.iloc[0]

            for col in common_cols:
                manual_value = float(row.get(col, 0)) if row.get(col) not in [None, ""] else 0
                calc_value = float(calc_row_data.get(col, 0)) if calc_row_data.get(col) not in [None, ""] else 0

                if manual_value != calc_value:
                    mismatches.append(col)
                    mismatch_details.append(f"{col} → Manual={manual_value} VSCode={calc_value}")
        else:
            # Site not found in calculated data
            for col in common_cols:
                manual_value = float(row.get(col, 0)) if row.get(col) not in [None, ""] else 0
                if manual_value != 0:
                    mismatches.append(col)
                    mismatch_details.append(f"{col} → Manual={manual_value} VSCode=0")

        # Build summary
        bom_df.at[idx, "Mismatch_Columns"] = ", ".join(mismatches) if mismatches else ""
        bom_df.at[idx, "Mismatch_Summary"] = "\n".join(mismatch_details) if mismatch_details else ""
        bom_df.at[idx, "Total_Mismatch_Count"] = len(mismatches)

        # Set status
        if len(mismatches) == 0:
            status = "MATCH"
            remark = "✅ MATCH"
        else:
            status = "MISMATCH"
            remark = "❌ MISMATCH - Semantic ItemKey validation detected differences."

        bom_df.at[idx, "Compare_Status"] = status
        bom_df.at[idx, "Compare_Remark"] = remark

    return bom_df


def format_excel(output_path: Path) -> None:
    wb = load_workbook(output_path)
    ws = wb.active

    # Colors for MATCH vs MISMATCH
    match_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    mismatch_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)

    # Find column indices for mismatch columns
    status_col = None
    mismatch_cols_idx = None
    mismatch_summary_idx = None
    count_idx = None

    for cell in ws[1]:
        if cell.value == "Compare_Status":
            status_col = cell.column
        elif cell.value == "Mismatch_Columns":
            mismatch_cols_idx = cell.column
        elif cell.value == "Mismatch_Summary":
            mismatch_summary_idx = cell.column
        elif cell.value == "Total_Mismatch_Count":
            count_idx = cell.column

    if status_col:
        for row in range(2, ws.max_row + 1):
            status_value = ws.cell(row=row, column=status_col).value

            if status_value == "MATCH":
                fill = match_fill
            else:
                fill = mismatch_fill

            # Apply fill to mismatch columns only
            for col_idx in [mismatch_cols_idx, mismatch_summary_idx, count_idx]:
                if col_idx:
                    cell_obj = ws.cell(row=row, column=col_idx)
                    cell_obj.fill = fill
                    cell_obj.font = white_font

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        adjusted_width = min(length + 3, 80)
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)


def main() -> None:
    print("=" * 60)
    print("BOM LOGIC COMPARISON")
    print("=" * 60)
    print()

    calc_df, bom_df, bom_site_col = load_files()
    compared_df = build_comparison_df(calc_df, bom_df, bom_site_col)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    compared_df.to_excel(OUTPUT_FILE, index=False)

    format_excel(OUTPUT_FILE)

    print()
    print("=" * 60)
    print("BOM LOGIC COMPARISON COMPLETE")

    print(f"Output: {OUTPUT_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()
