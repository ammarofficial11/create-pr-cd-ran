import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment

OPTIONAL_THRESHOLD_KEYS = [
    "AAU",
    "Antenna",
    "Battery",
    "BBU",
    "Bracket",
    "Cabinet",
    "Power Module",
    "Combiner",
    "PadPower",
    "Post",
    "RRU",
    "RRU Cage",
    "Security Bar",
    "Feeder 1/2",
    "Feeder 1 5/8",
    "Feeder 7/8",
]

OUTPUT_FILE = Path(__file__).resolve().parents[1] / "output" / "GeneralItem_Debug.xlsx"


def load_json_file(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def load_epms_site_map(epms_path: Path) -> dict:
    df = pd.read_excel(epms_path, sheet_name="data", header=3, engine="openpyxl")
    site_map = {}

    for _, row in df.iterrows():
        site_code = row.get("customer site code")
        if pd.isna(site_code):
            continue

        site_code = str(site_code).strip()
        site_map[site_code] = {
            "Region": str(row.get("region", "")).strip(),
            "ProvinceState": str(row.get("Province/State", "")).strip(),
            "City": str(row.get("City", "")).strip(),
            "SiteName": str(row.get("customer site name", "")).strip(),
        }

    return site_map


def build_analysis_dataframe(calculated: dict, general: dict, epms_map: dict) -> pd.DataFrame:
    rows = []

    for site_code, site_data in calculated.items():
        metadata = site_data.get("metadata", {}) or {}
        epms_lookup = epms_map.get(site_code, {})

        site_name = metadata.get("site_name") or epms_lookup.get("SiteName", "")
        region = metadata.get("region") or epms_lookup.get("Region", "")
        province_state = epms_lookup.get("ProvinceState", "")
        city = epms_lookup.get("City", "")

        general_rows = general.get(site_code, {}).get("pr_lines", [])
        general_item_count = len(general_rows)
        general_item_names = "\n".join(
            str(row.get("LineItemText", "")).strip()
            for row in general_rows
            if row.get("LineItemText") is not None
        )

        row = {
            "SiteCode": site_code,
            "SiteName": site_name,
            "Region": region,
            "ProvinceState": province_state,
            "City": city,
            "GeneralItemCount": general_item_count,
            "GeneralItemNames": general_item_names,
        }

        quantities = site_data.get("quantities", {}) or {}
        for key in OPTIONAL_THRESHOLD_KEYS:
            row[key] = quantities.get(key, 0) if quantities.get(key, 0) is not None else 0

        row["OptionalThresholdQty"] = sum(row[key] for key in OPTIONAL_THRESHOLD_KEYS)
        row["HasGeneralItems"] = "YES" if general_item_count > 0 else "NO"

        rows.append(row)

    df = pd.DataFrame(rows)
    df["GeneralItemCount"] = df["GeneralItemCount"].astype(int)
    df["OptionalThresholdQty"] = df["OptionalThresholdQty"].astype(float)

    df["__sort_has"] = df["HasGeneralItems"].map({"YES": 0, "NO": 1})
    df.sort_values(["__sort_has", "SiteCode"], inplace=True)
    df.drop(columns=["__sort_has"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    return df


def export_to_excel(df: pd.DataFrame, output_path: Path) -> None:
    sheet_name = "GeneralItemAnalysis"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_cells = next(worksheet.iter_rows(min_row=1, max_row=1))
        column_widths = {}

        for cell in header_cells:
            column_widths[cell.column_letter] = len(str(cell.value)) + 2

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                if cell.value is None:
                    continue
                cell_value = str(cell.value)
                width = len(cell_value) + 2
                if width > column_widths.get(cell.column_letter, 0):
                    column_widths[cell.column_letter] = min(width, 70)

        for column_letter, width in column_widths.items():
            worksheet.column_dimensions[column_letter].width = width

        name_col_index = df.columns.get_loc("GeneralItemNames") + 1
        for col_cells in worksheet.iter_cols(min_col=name_col_index, max_col=name_col_index, min_row=2, max_row=worksheet.max_row):
            for cell in col_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def print_summary(df: pd.DataFrame) -> None:
    total_sites = len(df)
    sites_with_general = int((df["GeneralItemCount"] > 0).sum())
    sites_without_general = int(total_sites - sites_with_general)
    average_threshold = df["OptionalThresholdQty"].mean()

    print(f"Total Sites: {total_sites}")
    print(f"Sites With General Items: {sites_with_general}")
    print(f"Sites Without General Items: {sites_without_general}")
    print(f"Average OptionalThresholdQty: {average_threshold:.2f}")


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    calculated_path = repo_root / "output" / "simple_calculated.json"
    general_path = repo_root / "output" / "general_pr_output.json"
    epms_path = repo_root / "input" / "EPMS.xlsx"

    calculated = load_json_file(calculated_path)
    general = load_json_file(general_path)
    epms_map = load_epms_site_map(epms_path)

    df = build_analysis_dataframe(calculated, general, epms_map)
    export_to_excel(df, OUTPUT_FILE)
    print_summary(df)
    print(f"Debug export created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
