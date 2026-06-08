import argparse
import json
import sys
from pathlib import Path

repo_root = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(repo_root))

import pandas as pd
from openpyxl.utils import get_column_letter

from src.simple_pr_generator import (
    _is_truthy_project_flag,
    _matches_location,
    _normalize_text,
    load_general_item_config,
)

OPTIONAL_THRESHOLD_KEYS = [
    "AAU",
    "Antenna",
    "Battery",
    "BBU",
    "Bracket",
    "Cabinet",
    "Power Module",
    "Combiner",
    "PadPower",
    "Post",
    "RRU",
    "RRU Cage",
    "Security Bar",
    "Feeder 1/2",
    "Feeder 1 5/8",
    "Feeder 7/8",
]

OUTPUT_FILE = Path(__file__).resolve().parents[1] / "output" / "GeneralItem_RejectionAnalysis.xlsx"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a detailed General Item rejection analysis report."
    )
    parser.add_argument(
        "--selected-project",
        dest="selected_project",
        help="Selected General DU project to analyze.",
        required=True,
    )
    return parser.parse_args()


def load_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def load_epms_map(path: Path) -> dict:
    df = pd.read_excel(path, sheet_name="data", header=3, engine="openpyxl")
    site_map = {}

    for _, row in df.iterrows():
        site_code = row.get("customer site code")
        if pd.isna(site_code):
            continue
        site_code = str(site_code).strip()
        site_map[site_code] = {
            "Region": str(row.get("region", "")).strip(),
            "ProvinceState": str(row.get("Province/State", "")).strip(),
            "City": str(row.get("City", "")).strip(),
            "SiteName": str(row.get("customer site name", "")).strip(),
        }

    return site_map


def get_config_region(region_value: str) -> str | None:
    region = _normalize_text(region_value)
    if region == "central":
        return "Central"
    if region == "sabah":
        return "Sabah"
    if region == "sarawak":
        return "Sarawak"
    if region in {"northern", "north"}:
        return "Northern"
    return None


def normalize_project_value(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def build_rejection_rows(
    calculated: dict,
    selected_project: str,
    general_config: dict,
    em_transport_map: dict,
    epms_map: dict,
) -> pd.DataFrame:
    rows = []

    for site_code, site_data in calculated.items():
        metadata = site_data.get("metadata", {}) or {}
        site_region = metadata.get("region", "")
        config_region = get_config_region(site_region)
        site_name = metadata.get("site_name", "")
        site_code_text = metadata.get("site_code", site_code)

        epms_site = epms_map.get(site_code, {})
        site_name_display = site_name or epms_site.get("SiteName", "")
        region_display = site_region or epms_site.get("Region", "")
        province_state = epms_site.get("ProvinceState", "")
        city = epms_site.get("City", "")

        quantities = site_data.get("quantities", {}) or {}
        optional_threshold_qty = sum(
            float(quantities.get(key, 0) or 0)
            for key in OPTIONAL_THRESHOLD_KEYS
        )

        if not config_region:
            continue

        config_rows = general_config.get(config_region, [])
        seen_material_codes = set()

        site_name_lower = _normalize_text(site_name_display)
        site_code_lower = _normalize_text(site_code_text)

        for row in config_rows:
            item_name = row.get("item_name", "")
            location = row.get("location", "")
            project_value_raw = row["project_flags"].get(selected_project)
            project_value = normalize_project_value(project_value_raw)

            decision = "ACCEPTED"
            reject_reason = ""

            if project_value == "":
                decision = "REJECTED"
                reject_reason = "Project flag blank"
            elif project_value.lower() == "optional":
                if optional_threshold_qty <= 6:
                    decision = "REJECTED"
                    reject_reason = "Optional threshold <= 6"
            elif not _is_truthy_project_flag(project_value_raw):
                decision = "REJECTED"
                reject_reason = "Project flag blank"

            if decision == "ACCEPTED":
                if config_region == "Central":
                    match = True
                else:
                    match = _matches_location(
                        site_code_text,
                        location,
                        config_region,
                        province_state,
                        city,
                        em_transport_map,
                    )
                if not match:
                    if not location:
                        reject_reason = "Location blank"
                    else:
                        location_value = _normalize_text(location)
                        matched_boq = False
                        for key, mapped_area in em_transport_map.items():
                            if key in site_name_lower or key in site_code_lower:
                                if _normalize_text(mapped_area) != location_value:
                                    matched_boq = True
                                    break
                        if matched_boq:
                            reject_reason = "BOQ Area mismatch"
                        elif config_region in {"Sabah", "Sarawak"}:
                            reject_reason = "City mismatch"
                        elif config_region == "Northern":
                            reject_reason = "Province mismatch"
                        else:
                            reject_reason = "Region mismatch"
                    decision = "REJECTED"

            item_code = row.get("item_code")
            if decision == "ACCEPTED":
                if item_code in seen_material_codes:
                    decision = "REJECTED"
                    reject_reason = "Unknown"
                else:
                    seen_material_codes.add(item_code)

            rows.append({
                "SiteCode": site_code,
                "SiteName": site_name_display,
                "Region": region_display,
                "ProvinceState": province_state,
                "City": city,
                "GeneralItemName": item_name,
                "GeneralItemRegionSheet": config_region,
                "GeneralItemLocation": location,
                "ProjectValue": project_value,
                "OptionalThresholdQty": optional_threshold_qty,
                "Decision": decision,
                "RejectReason": reject_reason,
            })

    df = pd.DataFrame(rows)
    df["OptionalThresholdQty"] = df["OptionalThresholdQty"].astype(float)
    return df


def create_summary_sheet(workbook, df: pd.DataFrame) -> None:
    summary = [
        ["Metric", "Value"],
        ["Accepted", int((df["Decision"] == "ACCEPTED").sum())],
        ["Rejected", int((df["Decision"] == "REJECTED").sum())],
        ["Rejected by Project", int((df["RejectReason"] == "Project flag blank").sum())],
        ["Rejected by Optional", int((df["RejectReason"] == "Optional threshold <= 6").sum())],
        ["Rejected by Region", int((df["RejectReason"] == "Region mismatch").sum())],
        ["Rejected by Province", int((df["RejectReason"] == "Province mismatch").sum())],
        ["Rejected by City", int((df["RejectReason"] == "City mismatch").sum())],
        ["Rejected by BOQ Area", int((df["RejectReason"] == "BOQ Area mismatch").sum())],
    ]

    from openpyxl import Workbook

    ws = workbook.create_sheet("Summary")
    for row in summary:
        ws.append(row)

    for col in range(1, 3):
        max_length = 0
        for cell in ws[get_column_letter(col)]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = max_length + 2


def export_report(df: pd.DataFrame, output_path: Path) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="RejectionAnalysis", index=False)
        workbook = writer.book
        worksheet = writer.sheets["RejectionAnalysis"]
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = "A2"

        for idx, column in enumerate(df.columns, start=1):
            max_length = len(column) + 2
            for cell in worksheet[get_column_letter(idx)]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)) + 2)
            worksheet.column_dimensions[get_column_letter(idx)].width = min(max_length, 70)

        create_summary_sheet(workbook, df)


def main() -> None:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    calculated_path = repo_root / "output" / "simple_calculated.json"
    epms_path = repo_root / "input" / "EPMS.xlsx"

    calculated = load_json(calculated_path)
    general_config, em_transport_map = load_general_item_config()
    epms_map = load_epms_map(epms_path)

    df = build_rejection_rows(
        calculated,
        args.selected_project,
        general_config,
        em_transport_map,
        epms_map,
    )
    export_report(df, OUTPUT_FILE)

    print(f"Debug export created: {OUTPUT_FILE}")
    print(f"Rows analyzed: {len(df)}")


if __name__ == "__main__":
    main()
